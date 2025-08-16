from apiclient import discovery
from httplib2 import Http
from oauth2client import client, file, tools
import json
import os
from openpyxl import Workbook, load_workbook
import time
import pandas as pd
from datetime import datetime
from tqdm import tqdm


#------------------------
# authen and author
#------------------------
def authen_and_author(DISCOVERY_DOC, SCOPES, authentication_dir):
    store = file.Storage(os.path.join(authentication_dir, "token.json"))
    creds = None

    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets(os.path.join(authentication_dir, "client_secrets.json"), SCOPES)
        creds = tools.run_flow(flow, store)

    form_service = discovery.build(
        "forms",
        "v1",
        http=creds.authorize(Http()),
        discoveryServiceUrl=DISCOVERY_DOC,
        static_discovery=False,
    )

    return form_service


#------------------------
# file excel tổng quan số sinh viên lầm khảo sát và danh sách chi tiết sinh viên
#------------------------
def total_do_survey(form_service, files_survey, files_response, name_file = "Tổng quan sinh viên thực hiện khảo sát.xlsx"):
      wb_survey = load_workbook(files_survey)
      ws_files_survey = wb_survey.active

      list_student_survey = [] # lưu sinh viên thực hiện khảo sát [sv1,  sv2]
      number_do_survey = {} # lưu số lượng thực hiện khảo sát của từng nhóm {TB101: 2}

      time = datetime.now().date() # ngày thực hiện file thống kê khảo sát

      #TODO: lấy dữ liệu khảo sát
      for row in tqdm(ws_files_survey.iter_rows(min_row=2), total=ws_files_survey.max_row, desc="Đang lấy dữ liệu sinh viên thực hiện khảo sát"):
            form_info = form_service.forms().get(formId=row[3].value).execute()
            question_map = {} # danh sách câu hỏi khảo sát thuộc từng form khát sát {id_question: name_question}
            for item in form_info.get("items", []):
                  if "questionItem" in item:
                        question_id = item["questionItem"]["question"]["questionId"]
                        question_text = item["title"]
                        question_map[question_id] = question_text
            
            responses = form_service.forms().responses().list(formId=row[3].value).execute() # số lượng sinh viên thực hiện khảo sát của từng form
            number_do_survey[row[0].value] = len(responses.get("responses", [])) # lưu số lượng sinh viên khảo sát

            #TODO: Lấy thông tin sinh viên thực hiện khảo sát
            for response in responses.get("responses", []):
                  for question_id, answer_data in response.get("answers", {}).items():
                        if "Mã số sinh viên:" in question_map[question_id]:
                              list_student_survey.append(answer_data["textAnswers"]["answers"][0]["value"])

      #TODO: Kiểm tra để tạo mới file thống kê hoặc thống kê dựa trên file hiện có
      total_response_excel = os.path.join(files_response, name_file) # tên file thống kê sinh viên thực hiện khảo sát
      if os.path.exists(total_response_excel):
            wb_response = load_workbook(total_response_excel)
            ws_general = wb_response["Tổng quan"]
            ws_detail = wb_response["Chi tiết"]

            #TODO: Thống kê số lượng sinh viên thực hiện khảo sát theo ngày dựa trên file thống kê đã có
            sum_survey = 0
            max_col_general = ws_general.max_column
            max_row_general = ws_general.max_row
            ws_general.cell(1, max_col_general + 1).value = time
            for index, row in tqdm(enumerate(ws_general.iter_rows(min_row=2)), total=max_row_general, desc="Đang cập nhật dữ liệu số lượng sinh viên thực hiện khảo sát"):
                  if row[0].value in number_do_survey.keys():
                        ws_general.cell(index + 2, max_col_general + 1).value = number_do_survey[row[0].value]
                        sum_survey += int(number_do_survey[row[0].value])
            ws_general.cell(max_row_general, max_col_general + 1).value = sum_survey

            #TODO: Thống kê sinh viên thực hiện khảo sát mới theo ngày dựa trên file thống kê đã có
            existed_sv = [sv[0].value for sv in ws_detail.iter_rows(min_row=1)]
            for sv in tqdm(list_student_survey, total=len(list_student_survey), desc="Đang cập nhật sinh viên mới làm khảo sát"):
                  if sv not in existed_sv:
                        max_row_detail = ws_detail.max_row
                        ws_detail.cell(max_row_detail + 1, 1).value = sv
                        ws_detail.cell(max_row_detail + 1, 2).value = time

            wb_response.save(total_response_excel)
      else:
            wb_new_response = Workbook()
            if "Sheet" in wb_new_response:
                  wb_new_response.remove(wb_new_response["Sheet"])
            ws_new_general = wb_new_response.create_sheet("Tổng quan")
            ws_new_detail = wb_new_response.create_sheet("Chi tiết")
            ws_new_general.append(["Nhóm đăng ký", time])
            ws_new_detail.append(["Mã số sinh viên", "Ngày thêm"])

            sum_survey = 0
            for key, value in tqdm(number_do_survey.items(), total=len(number_do_survey), desc="Đang thêm mới dữ liệu khảo sát"):
                  ws_new_general.append([key, value])
                  sum_survey += int(value)
            ws_new_general.append(["Tổng sinh viên thực hiện thực hiện khảo sát", sum_survey])

            for sv in tqdm(list_student_survey, total=len(list_student_survey), desc="Đang thêm mới sinh viên thực hiện khảo sát"):
                  ws_new_detail.append([sv, time])

            wb_new_response.save(total_response_excel)


#------------------------
# Thống kê câu trả lời của sinh viên
#------------------------
def detail_forms_survey(form_service, files_survey, file_response, name_file = "Kết quả thực hiện khảo sát.xlsx"):
      wb_file_response = load_workbook(files_survey) # đọc file lưu các form khảo sát
      ws_file_response = wb_file_response.active

      answers_survey = [] # Lưu các câu trả lời 
      title_question = [] # Lưu câu hỏi tương ứng với câu trả lời

      for row in tqdm(ws_file_response.iter_rows(min_row=2), total=ws_file_response.max_row, desc="Đang lấy dữ liệu từng form"):
            form_info = form_service.forms().get(formId=row[3].value).execute()
            set_questions = [] # Lưu từng bộ câu hỏi của từng môn trong 1 form khảo sát [{a,b,c}, {a,b,c}]
            questions_temp = {} # mảng tạm để lưu bộ câu hỏi, sau đó thêm vào set_question
            id_question = '' # lưu id của câu hỏi mã số sinh viên (vì form khảo sát gồm nhiều môn, cần tách riêng mssv cho từng môn)
            name_question = '' # lưu giá trị của câu hỏi mã số sinh viên
            index_question = 0 # đánh chỉ mục tất cả câu hỏi trong form

            #TODO: Lưu từng bộ câu hỏi của từng form, tương ứng với từng bộ câu hỏi sẽ có mssv
            for item in form_info.get("items", []):
                  if 'questionItem' in item:
                        if index_question == 0: # index = 0 chứa mssv
                              id_question =  item['questionItem']['question']['questionId']
                              name_question = item["title"]
                        else:
                              questions_temp[item['questionItem']['question']['questionId']] = item["title"]
                        if len(questions_temp) == 32: # một bộ câu hỏi gồm 32 câu
                              questions_temp[id_question] = name_question
                              set_questions.append(questions_temp)
                              questions_temp = {}
                        index_question += 1
            
            #TODO: Thực hiện lưu câu trả lời của từng form
            responses = form_service.forms().responses().list(formId=row[3].value).execute()
            for response in responses.get("responses", []):
                  response_temp = {} # mảng tạm lưu tất cả câu trả lời của form {id câu hỏi: câu trả lời}
                  for question_id, answer_data in response.get("answers", {}).items():
                        response_temp[question_id] = answer_data["textAnswers"]["answers"][0]["value"]
                  for questions in set_questions:
                        save_response = [] # mảng tạm lưu câu trả lời theo bộ 32 câu hỏi 
                        get_unit = '' # lưu thông tin đơn vị liên kết (được lấy từ câu hỏi)
                        for key, value_question in questions.items():
                              if "thông tin môn học" in value_question: # Tách chuỗi đơn vị liên kết vd: (BT - Bến tre) Thông tin môn học:
                                    unit = value_question.replace("thông tin môn học", "")
                                    unit = unit.replace("(", "")
                                    unit = unit.replace(")", "")
                                    unit = unit.replace(":", "")
                                    get_unit = unit
                              else: # những câu hỏi nào không chứa đơn vị liên kết, sẽ được lưu vào mảng để lấy tiêu đề câu hỏi
                                    if value_question not in title_question:
                                          title_question.append(value_question)
                              if key in response_temp.keys():
                                    save_response.append(response_temp[key])
                              else:
                                    save_response.append("")
                        save_response.append(get_unit)
                        answers_survey.append(save_response)
      
      detail_response_excel = os.path.join(file_response, name_file)
      wb_detail = Workbook()
      if "Sheet" in wb_detail:
            wb_detail.remove(wb_detail["Sheet"])
      ws_detail = wb_detail.create_sheet("Chi tiết")
      title_question.append("Địa phương")
      title_question.insert(0, "Thông tin môn học")
      ws_detail.append(list(title_question))
      for answer in answers_survey:
            ws_detail.append(answer)
      wb_detail.save(detail_response_excel)

      
def main():
    # TODO: scope of form
      SCOPES = [
            "https://www.googleapis.com/auth/forms.responses.readonly",
            "https://www.googleapis.com/auth/forms.body"]
      DISCOVERY_DOC = "https://forms.googleapis.com/$discovery/rest?version=v1"

      root_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
      #authentication_dir = os.path.join(root_folder, "authentication", "duy.nguyen2@oude.edu.vn")
      authentication_dir = os.path.join(root_folder, "authentication")
      files_dir = os.path.join(root_folder, "files")
      export_dir = os.path.join(root_folder, "export_file")
      files_survey = os.path.join(export_dir, "242", "log.xlsx")
      files_response = os.path.join(root_folder, "responses", "242")
      form_service = authen_and_author(DISCOVERY_DOC, SCOPES, authentication_dir) # authen and author

      detail_forms_survey(form_service, files_survey, files_response)

      #total_do_survey(form_service, files_survey, files_response)

if __name__ == "__main__":
      main()




