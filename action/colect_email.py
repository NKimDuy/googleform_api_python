from apiclient import discovery
from httplib2 import Http
from oauth2client import client, file, tools
import os
from openpyxl import load_workbook

#scope để tạo form
SCOPES = "https://www.googleapis.com/auth/forms.body"
DISCOVERY_DOC = "https://forms.googleapis.com/$discovery/rest?version=v1"

# scope để lấy dữ liệu từ form
# SCOPES = "https://www.googleapis.com/auth/forms.responses.readonly"
# DISCOVERY_DOC = "https://forms.googleapis.com/$discovery/rest?version=v1"


root_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
authentication_dir = os.path.join(root_folder, "authentication")
export_file_dir = os.path.join(root_folder, "export_file")
files_dir = os.path.join(root_folder, "files")
files_subject = os.path.join(files_dir, "241_ds_mon_hoc.xlsx")

store = file.Storage(os.path.join(authentication_dir, "token.json"))
creds = None
if not creds or creds.invalid:
  flow = client.flow_from_clientsecrets(os.path.join(authentication_dir, "client_secrets.json"), SCOPES)
  creds = tools.run_flow(flow, store)

form_service = discovery.build(
    "forms",
    "v1",
    http=creds.authorize(Http()),
    discoveryServiceUrl=DISCOVERY_DOC,
    static_discovery=False,
)


wb = load_workbook(files_subject)
ws = wb.active

for i in range(2, ws.max_row + 1):
    form_id = ws.cell(i, 31).value

    UPDATE_EMAIL_SETTINGS = {
        "requests": [
            {
                "updateSettings": {
                    "settings": {
                        "quizSettings": {
                            "isQuiz": False
                        },
                        "emailCollectionType": "VERIFIED"
                    },
                    "updateMask": "emailCollectionType"
                }
            }
        ]
    }

    response = form_service.forms().batchUpdate(formId=form_id, body=UPDATE_EMAIL_SETTINGS).execute()

    print(f"Đã cập nhật thu thập email cho form có id là: {form_id}")


