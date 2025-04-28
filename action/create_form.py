from apiclient import discovery
from httplib2 import Http
from oauth2client import client, file, tools
import os
from openpyxl import load_workbook
import time


# Xác định phạm vi của form
SCOPES = "https://www.googleapis.com/auth/forms.body"
DISCOVERY_DOC = "https://forms.googleapis.com/$discovery/rest?version=v1"


# đường dẫn đến các folder
root_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
authentication_dir = os.path.join(root_folder, "authentication")
export_file_dir = os.path.join(root_folder, "export_file")
files_dir = os.path.join(root_folder, "files")
files_subject = os.path.join(files_dir, "241_ds_mon_hoc.xlsx")


# Kiểm tra xác thực người dùng
store = file.Storage(os.path.join(authentication_dir, "token.json"))
creds = None
if not creds or creds.invalid:
  flow = client.flow_from_clientsecrets(os.path.join(authentication_dir, "client_secrets.json"), SCOPES)
  creds = tools.run_flow(flow, store)

form_service = discovery.build(
    "forms",
    "v1",
    http=creds.authorize(Http()),
    discoveryServiceUrl=DISCOVERY_DOC,
    static_discovery=False,
)


# Đọc file excel
wb = load_workbook(files_subject)
ws = wb.active

#root_url = "list_form"

# Xác định vị trí bắt đầu tạo form
begin_create = 2
for i in range(2, ws.max_row + 1):
   if ws.cell(i, 31).value is not None:
      begin_create += 1


print(f"Bắt đầu chạy từ số thứ tự là: {begin_create - 1}")

for i in range(begin_create, ws.max_row + 1):
    id_semester = "241"
    id_subject = ws.cell(i, 1).value
    name_subject = ws.cell(i, 2).value
    id_class = ws.cell(i, 3).value
    id_group = ws.cell(i, 8).value
    id_link_unit = ws.cell(i, 14).value
    name_link_unit = ws.cell(i, 15).value
    id_techer = ws.cell(i, 10).value
    name_teacher = ws.cell(i, 11).value
    name_manager = ws.cell(i, 30).value

    if not os.path.exists(os.path.join(export_file_dir, name_manager)):
        os.makedirs(os.path.join(export_file_dir, name_manager))
    link_manager = os.path.join(export_file_dir, name_manager)
    if not os.path.exists(os.path.join(link_manager, id_link_unit, name_link_unit)):
        os.makedirs(os.path.join(link_manager, id_link_unit, name_link_unit))
    link_unit = os.path.join(link_manager, id_link_unit, name_link_unit)
    if not os.path.exists(os.path.join(link_unit, id_class)):
        os.makedirs(os.path.join(link_unit, id_class))
    
    # Tạo thông tin forn ban đầu (chưa có thông tin câu hỏi)
    SURVEY_FORM = {
    "info": {
        "title": (
        "PHIẾU KHẢO SÁT Ý KIẾN SINH VIÊN "
        "VỀ HOẠT ĐỘNG GIẢNG DẠY CỦA SINH VIÊN "
        "Học kỳ 1 Năm học 2024 - 2025"
        ),
        "documentTitle": " - ".join([id_semester, id_subject, id_group, id_techer])
    }
    }

    ADD_INFO = {
    "requests": [
        {
        "updateFormInfo": {
            "info": {
            "description": (
                "Nhằm tăng cường tinh thần trách nhiệm của người học với quyền lợi, nghĩa vụ học tập, rèn luyện của bản thân: "
                "tạo điều kiện để người học được phản ánh tâm tư, nguyện vọng, được thể hiện chính kiến, "
                "nhà trường thực hiện khảo sát ý kiến sinh viên về hoạt động giảng dạy của giảng viên. "
                "Các bạn sinh viên vui lòng dành thười gian trả lời những câu hỏi dưới đây:"
            )
            },
            "updateMask": "description",
        }
        }
    ]
    }

# Thêm các câu trả lời vào form
    NEW_QUESTION = {
        "requests": [
            {
            "createItem": {
                "item": {
                "title": "PHẦN 1: THÔNG TIN CHUNG",
                "pageBreakItem": {}
                },
                "location": {"index": 0},
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Tên môn học:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {'value': " - ".join([id_subject, name_subject])}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 1},
            }
            }, 
            {
            "createItem": {
                "item": {
                "title": "Tên giảng viên:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {'value': " - ".join([id_techer, name_teacher])}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 2},
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Tên địa phương:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {'value': " - ".join([id_group, name_link_unit])}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 3},
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Mã số sinh viên:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "textQuestion": {}
                    }
                }
                },
                "location": {"index": 4}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "A: CHUẨN BỊ CHO MÔN HỌC",
                "pageBreakItem": {}
                },
                "location": {"index": 5}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Giảng viên (GV) giớ thiệu đề cương chi tiết và chuẩn đầu ra (CĐR) của môn học đầy đủ,rõ ràng trước khi bắt đầu môn học:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 6}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "GV giải thích phương pháp kiểm tra, đánh giá rõ ràng (thời điểm, nội dung, phương pháp kiểm tra, đánh giá) nhằm giúp sinh viên (SV) đạt được chuẩn đầu ra:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 7}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "GV giới thiệu nguồn tài liệu tham khảo:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 8}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Tài liệu được phát kịp thời cho môn học:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 9}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "B. PHƯƠNG PHÁP GIẢNG DẠY CỦA GIẢNG VIÊN",
                "pageBreakItem": {}
                },
                "location": {"index": 10}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Phương pháp truyền đạt rõ ràng, dễ hiểu nhằm giúp SV đạt được chuẩn đầu ra:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 11}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Cách thức giảng dạy tạo hứng thú học tập cho người học:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 12}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Tạo điều kiện để SV tham gia tích cực vào các hoạt động trong tiết học:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 13}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Nêu vấn đề để SV tham gia tích cực vào các hoạt động trong tiết học:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 14}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Hướng dẫn sinh viên cách tự học, tự nghiên cứu ngoài giờ học:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 15}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Sử dụng hiệu quả các phương tiện dạy học (máy chiếu, internet...):",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 16}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "GV quan tâm đến việc tiếp thu bài giảng của sinh viên:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 17}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Phương pháp truyền đạt rõ ràng, dễ hiểu nhằm giúp SV đạt được chuẩn đầu ra:",
                "pageBreakItem": {}
                },
                "location": {"index": 18}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Nội dung bài giảng được trình bày đầy đủ theo đề cương chi tiết môn học:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 19}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Bổ xung, cập nhật những vấn đề mới bên ngoài nội dung của giáo trình:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 20}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Nội dung môn học được cập nhật phù hợp với thực tiễn:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 21}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Phương pháp truyền đạt rõ ràng, dễ hiểu nhằm giúp SV đạt được chuẩn đầu ra:",
                "pageBreakItem": {}
                },
                "location": {"index": 22}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Thực hiện nghiêm túc giờ giấc giảng dạy, sử dụng hiệu quả thời gian lên lớp:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 23}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Nhiệt tình và có trách nhiệm trong giảng dạy:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 24}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Thể hiện tính chuẩn mực tác phong nhà giáo: trang phục, lời nới, cử chỉ:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 25}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Có thái độ tôn trọng người học:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 26}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "GV có sử dụng hiệu quả công nghệ hỗ trợ giảng dạy và học tập (Hệ thống quản lý học tập LMS):",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 27}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "GV theo đúng thời khóa biểu nhà trường đã đề ra:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 28}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "GV giảng dạy theo đúng tài liệu nhà trường đã cung cấp:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 29}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Thời lượng hướng dẫn/giảng dạy của môn học là phù hợp:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 30}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "E. HOẠT ĐỘNG KIỂM TRA, ĐÁNH GIÁ QUÁ TRÌNH HỌC TẬP",
                "pageBreakItem": {}
                },
                "location": {"index": 31}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Kết quả kiểm tra giữa kỳ được GV công bố trước khi kết thúc môn học:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 32}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "GV sử dụng nhiều hình thức kiểm tra, đánh giá để tăng độ chính xác, tin cậy, tính giá trị trong đánh giá và đáp ứng CĐR:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 33}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "GV đánh giá công bằng và phản ánh đúng năng lực của SV theo chuẩn đầu ra (CĐR):",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 34}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Nội dung kiểm tra phù hợp với nội dung giảng dạy và CĐR:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 35}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Tài liệu học tập được cung cấp đúng với thông tin ghi trên đề cương môn học:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Hoàn toàn không đồng ý"},
                        {"value": "Không đồng ý"},
                        {"value": "Phân vân"},
                        {"value": "Đồng ý"},
                        {"value": "Hoàn toàn đồng ý"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 36}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "F. HOẠT ĐỘNG KIỂM TRA, ĐÁNH GIÁ QUÁ TRÌNH HỌC TẬP",
                "pageBreakItem": {}
                },
                "location": {"index": 37}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Anh/chị cho biết mức độ hài lòng về chất lượng giảng dạy của giảng viên:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Rất không hài lòng"},
                        {"value": "Không hài lòng"},
                        {"value": "Hài lòng trung bình"},
                        {"value": "Khá hài lòng"},
                        {"value": "Rất hài lòng"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 38}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Anh/chị cho biết mức độ hài lòng về hiệu quả giảng dạy của giảng viên:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Rất không hài lòng"},
                        {"value": "Không hài lòng"},
                        {"value": "Hài lòng trung bình"},
                        {"value": "Khá hài lòng"},
                        {"value": "Rất hài lòng"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 39}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Nhìn chung (tổng thể), Anh/Chị cho biết mức độ hài lòng về chất lượng & hiệu quả giảng dạy của giảng viên:",
                "questionItem": {
                    "question": {
                    "required": True,
                    "choiceQuestion": {
                        "type": "RADIO",
                        "options": [
                        {"value": "Rất không hài lòng"},
                        {"value": "Không hài lòng"},
                        {"value": "Hài lòng trung bình"},
                        {"value": "Khá hài lòng"},
                        {"value": "Rất hài lòng"}
                        ]
                    }
                    }
                }
                },
                "location": {"index": 40}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Tổng kết:",
                "pageBreakItem": {}
                },
                "location": {"index": 41}
            }
            },
            {
            "createItem": {
                "item": {
                "title": "Ý kiến khác (nếu có):",
                "questionItem": {
                    "question": {
                    "required": False,
                    "textQuestion": {}
                    }
                }
                },
                "location": {"index": 42}
            }
            },
            {
            "createItem": {
                "item": {
                   "title": "LƯU Ý:",
                   "description": (
                      "Sinh viên cần tiếp tục thực hiện đánh giá các môn học khác"
                      " cho đến khi đủ số môn học mà sinh viên đã đăng ký trong học kỳ này."
                   ),
                   "textItem": {}
                },
                "location": {"index": 43}
            }
            },
        ]
    }

    # Khởi tạo form trống
    result = form_service.forms().create(body=SURVEY_FORM).execute()


    # Thêm description vào form
    question_setting = (
        form_service.forms()
        .batchUpdate(formId=result["formId"], body=ADD_INFO)
        .execute()
    )


    # Thêm các câu hỏi vào form
    question_setting = (
        form_service.forms()
        .batchUpdate(formId=result["formId"], body=NEW_QUESTION)
        .execute()
    )

    #Hiển thị kết quả để kiểm tra

    form_id = result["formId"]
    name_file = "".join([id_subject, "-", id_group, "-", id_techer, ".txt"])
    with open(os.path.join(link_unit, name_file), "w", encoding="utf-8") as file:
      file.write(f"https://docs.google.com/forms/d/{form_id}/viewform")
    ws.cell(i, 31).value = form_id
    wb.save(files_subject)
    
    print(f" Đang tạo form có ID {i - 1}: {form_id}")
    if i == ws.max_row:
       print(f"Đã thêm thành công {i} dòng")
    
    time.sleep(12)