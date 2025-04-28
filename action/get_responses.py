from apiclient import discovery
from httplib2 import Http
from oauth2client import client, file, tools
import json
import os
from openpyxl import Workbook, load_workbook
import time

SCOPESE = [
      "https://www.googleapis.com/auth/forms.responses.readonly",
      "https://www.googleapis.com/auth/forms.body"]
DISCOVERY_DOC = "https://forms.googleapis.com/$discovery/rest?version=v1"

root_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
authentication_dir = os.path.join(root_folder, "authentication")
files_dir = os.path.join(root_folder, "files")
files_subject = os.path.join(files_dir, "get_response.xlsx")
files_241 = os.path.join(files_dir, "241_ds_mon_hoc.xlsx")

store = file.Storage(os.path.join(authentication_dir, "token.json"))
# creds = None
# if not creds or creds.invalid:
#   flow = client.flow_from_clientsecrets(os.path.join(authentication_dir, "client_secrets.json"), SCOPES_FORM)
#   creds = tools.run_flow(flow, store)

# service = discovery.build(
#     "forms",
#     "v1",
#     http=creds.authorize(Http()),
#     discoveryServiceUrl=DISCOVERY_DOC,
#     static_discovery=False,
# )

# Prints the responses of your specified form:
#form_id = "1RMu3NJGzC4kVqAQ2ln4w3MRH3BDuBvoygFndI9AE1mY"
# result = service.forms().responses().list(formId=form_id).execute()
# print(result)


# if not os.path.exists(files_subject):
#   wb = Workbook()
#   ws = wb.active
#   ws.title = "Sheet 1"

#   get_result = service.forms().get(formId=form_id).execute()
#   question_map = {}
#   create_title = []
#   for item in get_result.get("items", []):
#     if "questionItem" in item:
#       question_id = item["questionItem"]["question"]["questionId"]
#       question_text = item["title"]
#       question_map[question_id] = question_text
#       create_title.append(question_id)
    
#     #question_map[question_id] = question_text
#   ws.append(create_title)
#   wb.save(files_subject)
# else:
#   print(files_subject)
# time.sleep(15)

#   print(get_result)
#xuất dữ liệu vào file json
# with open("question.json", "w", encoding="utf-8") as json_file:
#     json.dump(get_result, json_file, ensure_ascii=False, indent=4)  # indent=4 giúp dễ đọc hơn


# lấy danh sách các câu hỏi
# Tạo một dictionary mapping giữa questionId và nội dung câu hỏi
# question_map = {}
# for item in form_info.get("items", []):
#     if "questionItem" in item:
#         question_id = item["questionItem"]["question"]["questionId"]
#         question_text = item["title"]
#         question_map[question_id] = question_text

# print(question_map)  # Kiểm tra danh sách câu hỏi


#ánh xạ câu trả lời vào câu hỏi
creds = None
if not creds or creds.invalid:
  flow = client.flow_from_clientsecrets(os.path.join(authentication_dir, "client_secrets.json"), SCOPESE)
  creds = tools.run_flow(flow, store)

service = discovery.build(
    "forms",
    "v1",
    http=creds.authorize(Http()),
    discoveryServiceUrl=DISCOVERY_DOC,
    static_discovery=False,
)



# ------- khu vực đang thao tác --------------------
wb = load_workbook(files_241)
ws = wb.active

overall = []

for i in range(2, ws.max_row + 1):
    form_info = service.forms().get(formId=ws.cell(i, 31).value).execute()
    question_map = {}
    for item in form_info.get("items", []):
        if "questionItem" in item:
            question_id = item["questionItem"]["question"]["questionId"]
            question_text = item["title"]
            question_map[question_id] = question_text

    count_responses_of_form = 0

    responses = service.forms().responses().list(formId=ws.cell(i, 31).value).execute()
    for response in responses.get("responses", []):
        test = {}
        test["Thời gian submit form"] = response["lastSubmittedTime"]
        test["email sinh viên"] = response["respondentEmail"]
        for question_id, answer_data in response.get("answers", {}).items():
            test[question_map[question_id]] = answer_data["textAnswers"]["answers"][0]["value"]
        overall.append(test)
        count_responses_of_form += 1
    print(f" -------------------Đang chạy from thứ {i - 1} và trong form có tổng cộng {count_responses_of_form} phản hồi -------------------")



all_questions = set()
for response in overall:
    all_questions.update(response.keys())

all_questions = sorted(all_questions)

wb = Workbook()
ws = wb.active
ws.title = "Sheet 1"
ws.append(all_questions)

for response in overall:
    row = [response.get(col, "") for col in all_questions]
    ws.append(row)
    print(f"Đã thêm: {",".join(map(str, row))}")
wb.save(files_subject)
#------------------------------------------------------------------------




    #   if not col_title:
    #       col_title.append(question_id)
    #   row.append(answer_data["textAnswers"]["answers"][0]["value"])
    # rows_value.append(row)
#print(rows_value)

# wb = Workbook()
# ws = wb.active
# ws.title = "Sheet 1"

# for item in overall:
#    ws.append(list(item.keys()))
#    ws.append(list(item.values()))

# wb.save(files_subject)
        

    #print(f"Phản hồi ID: {response['responseId']}")
    # for question_id, answer_data in response.get("answers", {}).items():
    #     question_text = question_map.get(question_id, "Câu hỏi không xác định")
    #     answer_value = answer_data["textAnswers"]["answers"][0]["value"]
    #     test1.append(question_id)
    #     test2.append(question_text)
    # if not col_title:

    #   col_title.append(response["answers"]["question_id"])

      # for question_id, answer_data in response.get("answers", {}).items():
      #   #question_text = question_map.get(question_id, "Câu hỏi không xác định")
      #   answer_value = answer_data["textAnswers"]["answers"][0]["value"]
#print(col_title)

#wb.save(files_subject)