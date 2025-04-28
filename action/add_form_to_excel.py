from apiclient import discovery
from httplib2 import Http
from oauth2client import client, file, tools
import os
from openpyxl import load_workbook
import time

root_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
authentication_dir = os.path.join(root_folder, "authentication")
export_file_dir = os.path.join(root_folder, "export_file")
files_dir = os.path.join(root_folder, "files")
files_subject = os.path.join(files_dir, "241_ds_mon_hoc.xlsx")

wb = load_workbook(files_subject)
ws = wb.active


wb = load_workbook(files_subject)
ws = wb.active

for i in range(2, ws.max_row + 1):
    form_id = ws.cell(i, 31).value
    ws.cell(i, 32).value = "".join(["https://docs.google.com/forms/d/", form_id, "/viewform"])
    wb.save(files_subject)

    print(f"đã thêm form {i - 1}")
