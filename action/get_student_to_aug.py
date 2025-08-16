from apiclient import discovery
from httplib2 import Http
from oauth2client import client, file, tools
import json
import os
from openpyxl import Workbook, load_workbook
import time
import pandas as pd
from datetime import datetime
from tqdm import tqdm
import random

def main():
    # TODO: scope of form
      root_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
      export_dir = os.path.join(root_folder, "export_file")
      files_survey = os.path.join(export_dir, "242", "log.xlsx")
      files_response = os.path.join(root_folder, "responses", "242")

      file_student = os.path.join(files_response, "random_sv.xlsx")
      wb_read = load_workbook(file_student)
      ws_read = wb_read.active

      dic_initial = dict()
      dic_choice_sv = dict()

      for row in ws_read.iter_rows():
           if row[0].value not in dic_initial.keys():
                  dic_initial[row[0].value] = []
      for row in ws_read.iter_rows():
            dic_initial[row[0].value].append(row[1].value)
      
      
      for key, arr_value in dic_initial.items():
            dic_choice_sv[key] = random.sample(arr_value, len(arr_value) // 2)
      
      wb_create = Workbook()
      ws_create = wb_create.active
      for key, arr_value in dic_choice_sv.items():
            for arr in arr_value:
                  ws_create.append([key, arr])
      export_file = os.path.join(files_response, "choiced_sv.xlsx")
      wb_create.save(export_file)

      
if __name__ == "__main__":
      main()
