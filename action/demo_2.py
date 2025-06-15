from apiclient import discovery
from httplib2 import Http
from oauth2client import client, file, tools
import os
from openpyxl import load_workbook, Workbook
import time


#------------------------
# authen and author
#------------------------
def authen_and_author(DISCOVERY_DOC, SCOPES, authentication_dir):
    store = file.Storage(os.path.join(authentication_dir, "token.json"))
    creds = None

    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets(os.path.join(authentication_dir, "client_secrets.json"), SCOPES)
        creds = tools.run_flow(flow, store)

    form_service = discovery.build(
        "forms",
        "v1",
        http=creds.authorize(Http()),
        discoveryServiceUrl=DISCOVERY_DOC,
        static_discovery=False,
    )

    return form_service


#------------------------
# form'name and info
#------------------------
def create_form_id(form_service, semester, group):
      SURVEY_FORM = {
            "info": {
            "title": (
            "PHIẾU KHẢO SÁT Ý KIẾN SINH VIÊN "
            "VỀ HOẠT ĐỘNG GIẢNG DẠY CỦA SINH VIÊN "
            "HỌC KỲ 2 NĂM HỌC 2024 - 2025"
            ),
            "documentTitle": " - ".join([semester, group])
            }
      }

      # TODO: saved form path to file
      result = form_service.forms().create(body=SURVEY_FORM).execute()
      form_id = result["formId"]
      return form_id


#------------------------
# add description form
#------------------------
def add_info_form(form_service, form_id):
      ADD_INFO = {
            "requests": [
            {
            "updateFormInfo": {
                  "info": {
                  "description": (
                        "Nhằm tăng cường tinh thần trách nhiệm của người học với quyền lợi, nghĩa vụ học tập, rèn luyện của bản thân: "
                        "tạo điều kiện để người học được phản ánh tâm tư, nguyện vọng, được thể hiện chính kiến, "
                        "nhà trường thực hiện khảo sát ý kiến sinh viên về hoạt động giảng dạy của giảng viên. "
                        "Các bạn sinh viên vui lòng dành thời gian trả lời những câu hỏi dưới đây:"
                  )
                  },
                  "updateMask": "description",
            }
            }
            ]
      }
      form_service.forms().batchUpdate(formId=form_id, body=ADD_INFO).execute()



#------------------------
# add header form
#------------------------
def add_header_form(form_service, form_id):
      NEW_HEADER = {
            "requests": [
                  {
                        "createItem": {
                              "item": {
                                    "title": "PHẦN 1: THÔNG TIN CHUNG",
                                    "textItem": {}
                              },
                              "location": {"index": 0},
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Mã số sinh viên:",
                                    "questionItem": {
                                          "question": {
                                                "required": True,
                                                "textQuestion": {}
                                          }
                                    }
                              },
                              "location": {"index": 1},
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "TIẾN HÀNH ĐÁNH GIÁ MÔN HỌC",
                                    "description": "Sinh viên vui lòng làm khảo sát hết tất cả các môn đã đăng ký",
                                    "pageBreakItem": {}
                              },
                              "location": {"index": 2},
                        }
                  }
            ]   
      }
      form_service.forms().batchUpdate(formId=form_id, body=NEW_HEADER).execute()


#------------------------
# add body form
#------------------------
def add_body_form(form_service, form_id, link_unit_dir, subject, teacher, location):
      NEW_QUESTION = {
            "requests": [
                  {
                        "createItem": {
                              "item": {
                                    "title": f"({link_unit_dir}) thông tin môn học: ",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": f"📌📌📌 ({subject}) ({teacher})"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 1}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Giảng viên (GV) giớ thiệu đề cương chi tiết và chuẩn đầu ra (CĐR) của môn học đầy đủ, rõ ràng trước khi bắt đầu môn học:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 2}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "GV giải thích phương pháp kiểm tra, đánh giá rõ ràng (thời điểm, nội dung, phương pháp kiểm tra, đánh giá) nhằm giúp sinh viên (SV) đạt được chuẩn đầu ra:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 3}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "GV giới thiệu nguồn tài liệu tham khảo:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 4}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Tài liệu được phát kịp thời cho môn học:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 5}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                              "title": "Phương pháp truyền đạt rõ ràng, dễ hiểu nhằm giúp SV đạt được chuẩn đầu ra:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 6}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Cách thức giảng dạy tạo hứng thú học tập cho người học:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 7}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Tạo điều kiện để SV tham gia tích cực vào các hoạt động trong tiết học:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 8}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Nêu vấn đề để SV tham gia tích cực vào các hoạt động trong tiết học:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 9}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Hướng dẫn sinh viên cách tự học, tự nghiên cứu ngoài giờ học:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 10}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Sử dụng hiệu quả các phương tiện dạy học (máy chiếu, internet...):",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 11}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "GV quan tâm đến việc tiếp thu bài giảng của sinh viên:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 12}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Nội dung bài giảng được trình bày đầy đủ theo đề cương chi tiết môn học:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 13}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Bổ xung, cập nhật những vấn đề mới bên ngoài nội dung của giáo trình:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 14}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Nội dung môn học được cập nhật phù hợp với thực tiễn:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 15}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Thực hiện nghiêm túc giờ giấc giảng dạy, sử dụng hiệu quả thời gian lên lớp:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 16}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Nhiệt tình và có trách nhiệm trong giảng dạy:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 17}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Thể hiện tính chuẩn mực tác phong nhà giáo: trang phục, lời nới, cử chỉ:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 18}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Có thái độ tôn trọng người học:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 19}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "GV có sử dụng hiệu quả công nghệ hỗ trợ giảng dạy và học tập (Hệ thống quản lý học tập LMS):",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 20}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "GV theo đúng thời khóa biểu nhà trường đã đề ra:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 21}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "GV giảng dạy theo đúng tài liệu nhà trường đã cung cấp:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 22}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Thời lượng hướng dẫn/giảng dạy của môn học là phù hợp:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 23}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Kết quả kiểm tra giữa kỳ được GV công bố trước khi kết thúc môn học:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 24}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "GV sử dụng nhiều hình thức kiểm tra, đánh giá để tăng độ chính xác, tin cậy, tính giá trị trong đánh giá và đáp ứng CĐR:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 25}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "GV đánh giá công bằng và phản ánh đúng năng lực của SV theo chuẩn đầu ra (CĐR):",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 26}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Nội dung kiểm tra phù hợp với nội dung giảng dạy và CĐR:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 27}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Tài liệu học tập được cung cấp đúng với thông tin ghi trên đề cương môn học:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Hoàn toàn không đồng ý"},
                                                            {"value": "Không đồng ý"},
                                                            {"value": "Phân vân"},
                                                            {"value": "Đồng ý"},
                                                            {"value": "Hoàn toàn đồng ý"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 28}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Anh/chị cho biết mức độ hài lòng về chất lượng giảng dạy của giảng viên:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Rất không hài lòng"},
                                                            {"value": "Không hài lòng"},
                                                            {"value": "Hài lòng trung bình"},
                                                            {"value": "Khá hài lòng"},
                                                            {"value": "Rất hài lòng"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 29}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Anh/chị cho biết mức độ hài lòng về hiệu quả giảng dạy của giảng viên:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Rất không hài lòng"},
                                                            {"value": "Không hài lòng"},
                                                            {"value": "Hài lòng trung bình"},
                                                            {"value": "Khá hài lòng"},
                                                            {"value": "Rất hài lòng"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 30}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Nhìn chung (tổng thể), Anh/Chị cho biết mức độ hài lòng về chất lượng & hiệu quả giảng dạy của giảng viên:",
                                    "questionItem": {
                                          "question": {
                                                "choiceQuestion": {
                                                      "type": "RADIO",
                                                      "options": [
                                                            {"value": "Rất không hài lòng"},
                                                            {"value": "Không hài lòng"},
                                                            {"value": "Hài lòng trung bình"},
                                                            {"value": "Khá hài lòng"},
                                                            {"value": "Rất hài lòng"}
                                                      ]
                                                }
                                          }
                                    }
                              },
                              "location": {"index": location + 31}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "Ý kiến khác (nếu có):",
                                    "questionItem": {
                                          "question": {
                                          "textQuestion": {}
                                          }
                                    }
                              },
                              "location": {"index": location + 32}
                        }
                  },
                  {
                        "createItem": {
                              "item": {
                                    "title": "ĐÁNH GIÁ MÔN HỌC",
                                    "pageBreakItem": {}
                              },
                              "location": {"index": location + 33}
                        }
                  }
            ]
      }
      form_service.forms().batchUpdate(formId=form_id, body=NEW_QUESTION).execute()


#------------------------
# form'name and info + add description form + add body form
#------------------------
def completed_form(grp, link_unit):

      form_id = "a"

      # TODO: saved form path to file
      name_file = "".join([grp, ".txt"])
      with open(os.path.join(link_unit, name_file), "w", encoding="utf-8") as file:
            file.write(f"https://docs.google.com/forms/d/{form_id}/viewform")
      return form_id


#------------------------
# loaded excel file, make dir and created log file
#------------------------
def read_file_and_create_form(form_service, export_dir, file, semester_dir):
      wb_subject = load_workbook(file)
      ws_subject = wb_subject.active

      # TODO: check log file exist or no
      log_file = os.path.join(export_dir, "log.xlsx")
      if not os.path.exists(log_file): # if log file not existed
            wb_create_log = Workbook()
            ws_create_log = wb_create_log.active
            ws_create_log.title = "log file"
            wb_create_log.save(log_file)
      wb_load_log = load_workbook(log_file) # if log file existed then created file
      ws_load_log = wb_load_log.active

      data_subject = {} # store subject
      data_teacher = {} # store teacher
      data_manager = {} # store name manager
      data_unit = {} # store unit
      for row in ws_subject.iter_rows(min_row=2, values_only=True):
            id_subject = row[0]
            name_subject = row[1]
            group = row[7]
            id_teacher = row[9] 
            name_teacher = row[10]
            name_manager = row[21]
            id_unit = row[13]
            name_unit = row[14]

            # TODO: read excel file, and attached to array
            if group not in data_subject:
                  data_subject[group] = []
                  data_teacher[group] = []
                  data_manager[group] = []
                  data_unit[group] = []
            data_subject[group].append(id_subject + "-" + name_subject)
            data_teacher[group].append(id_teacher + "-" + name_teacher)
            data_manager[group].append(name_manager)
            data_unit[group].append(id_unit + "-" + name_unit)

      # TODO: loop through each subject with corresponding instructor
      for grp in data_subject:
            name_manager_dir = list(set(data_manager[grp]))[0] # get unique manager 
            link_unit_dir = list(set(data_unit[grp]))[0] # get unique link unit

            # TODO: create or open semester dir
            if not os.path.exists(os.path.join(export_dir, semester_dir)):
                  os.makedirs(os.path.join(export_dir, semester_dir))
            link_semester = os.path.join(export_dir, semester_dir)

            # TODO: create or open manager dir
            if not os.path.exists(os.path.join(link_semester, name_manager_dir)):
                  os.makedirs(os.path.join(link_semester, name_manager_dir))
            link_manager = os.path.join(link_semester, name_manager_dir)
            
            # TODO: create or open unit dir
            if not os.path.exists(os.path.join(link_manager, link_unit_dir)):
                  os.makedirs(os.path.join(link_manager, link_unit_dir))
            link_unit = os.path.join(link_manager, link_unit_dir)
            
            # TODO: check if log file have or not have data
            has_data = any(row for row in ws_load_log.iter_rows(min_row=1, values_only=True) if any(row)) # return true, if at least one value is True 
            if has_data:
                  # TODO: check group exist in log file 
                  group_in_log = [row[0].value for row in ws_load_log.iter_rows(min_row=1)]
                  if grp not in group_in_log:

                        form_id = create_form_id(form_service, semester_dir, grp)
                        add_info_form(form_service, form_id)
                        add_header_form(form_service, form_id)
                        
                        location = 2
                        for i in range(len(data_subject[grp])):
                              print(f"{location} - {data_subject[grp][i]} - {data_teacher[grp][i]}")
                              add_body_form(form_service, form_id, link_unit_dir, data_subject[grp][i], data_teacher[grp][i], location)
                              location += 33

                        name_file = "".join([grp, ".txt"])
                        with open(os.path.join(link_unit, name_file), "w", encoding="utf-8") as file:
                              file.write(f"https://docs.google.com/forms/d/{form_id}/viewform")
                  
                        print(f"Bắt đầu ghi tiếp tục vào file từ dòng: {ws_load_log.max_row}: {grp} - {link_unit_dir} - {name_manager_dir} - {form_id}")
                        ws_load_log.append([grp, link_unit_dir, name_manager_dir, form_id])
                        wb_load_log.save(log_file)
            else:
                  form_id = create_form_id(form_service, semester_dir, grp)
                  add_info_form(form_service, form_id)
                  add_header_form(form_service, form_id)
                  
                  location = 2
                  for i in range(len(data_subject[grp])):
                        print(f"{location} - {data_subject[grp][i]} - {data_teacher[grp][i]}")
                        add_body_form(form_service, form_id, link_unit_dir, data_subject[grp][i], data_teacher[grp][i], location)
                        location += 33
                  
                  name_file = "".join([grp, ".txt"])
                  with open(os.path.join(link_unit, name_file), "w", encoding="utf-8") as file:
                        file.write(f"https://docs.google.com/forms/d/{form_id}/viewform")

                  print(f"Bắt đầu ghi mới vào file: {grp} - {link_unit_dir} - {name_manager_dir} - {form_id}")
                  ws_load_log.append([grp, link_unit_dir, name_manager_dir, form_id])
                  wb_load_log.save(log_file)
            

def main():
    # TODO: scope of form
    SCOPES = "https://www.googleapis.com/auth/forms.body"
    DISCOVERY_DOC = "https://forms.googleapis.com/$discovery/rest?version=v1"

    root_folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    authentication_dir = os.path.join(root_folder, "authentication", "duy.nguyen2@oude.edu.vn")
    export_dir = os.path.join(root_folder, "export_file")
    files_dir = os.path.join(root_folder, "files")
    files_subject = os.path.join(files_dir, "test.xlsx")

    form_service = authen_and_author(DISCOVERY_DOC, SCOPES, authentication_dir) # authen and author
    read_file_and_create_form(form_service, export_dir, files_subject, "242") # create form

if __name__ == "__main__":
      main()